# formula_detection.py
import openpyxl
import re

# ---------------- QUARTER TO MONTH MAPPING ----------------
quarter_months = {
    "1st Qtr": ["Jan", "Feb", "Mar"],
    "2nd Qtr": ["Apr", "May", "Jun"],
    "3rd Qtr": ["Jul", "Aug", "Sep"],
    "4th Qtr": ["Oct", "Nov", "Dec"]
}

# ---------------- NORMALIZE FORMULA ----------------
def normalize_formula(f):
    if f is None:
        return ""
    f = str(f).replace(" ", "").replace("=", "")
    f = re.sub(r'IFERROR\((.*?),""\)', r'\1', f)
    f = re.sub(r'IF\(AND\(ISNUMBER\(.*?\),ISNUMBER\(.*?\)\),(.*?),""\)', r'\1', f)
    if "SUM(" in f and "+" in f:
        f = f.replace("+", ",")
    return f

# ---------------- RECORD ERRORS ----------------
def record(errors_list, sheet, row, column, given, correct):
    errors_list.append({
        "sheet": sheet,
        "row": row,
        "column": column,
        "given_formula": given,
        "expected_logic": correct
    })

# ---------------- VALIDATE FILE ----------------
def analyze_formula_errors(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    errors = []

    # ---------------- YEAR TO DATE VALIDATION ----------------
    if "Year to Date" in wb.sheetnames:
        sheet = wb["Year to Date"]
        for row in range(4, sheet.max_row + 1):
            # ACTUAL
            val = normalize_formula(sheet[f"C{row}"].value)
            correct = f"SUM('1stQtr'!C{row},'2ndQtr'!C{row},'3rdQtr'!C{row},'4thQtr'!C{row})"
            if val and val != normalize_formula(correct):
                record(errors, "Year to Date", row, "Actual", val, correct)

            # EXPECTED
            val = normalize_formula(sheet[f"D{row}"].value)
            correct = f"SUM('1stQtr'!D{row},'2ndQtr'!D{row},'3rdQtr'!D{row},'4thQtr'!D{row})"
            if val and val != normalize_formula(correct):
                record(errors, "Year to Date", row, "Expected", val, correct)

            # DELTA
            val = normalize_formula(sheet[f"E{row}"].value)
            base = f"C{row}-D{row}"
            sum_qtr = f"SUM('1stQtr'!E{row},'2ndQtr'!E{row},'3rdQtr'!E{row},'4thQtr'!E{row})"
            if val and val not in [normalize_formula(base), normalize_formula(sum_qtr)]:
                record(errors, "Year to Date", row, "Delta", val, base)

            # %
            val = normalize_formula(sheet[f"F{row}"].value)
            base = f"E{row}/D{row}"
            if val and base not in val:
                record(errors, "Year to Date", row, "Expected %", val, base)

            # FORECASTED
            val = normalize_formula(sheet[f"G{row}"].value)
            correct = f"SUM('1stQtr'!G{row},'2ndQtr'!G{row},'3rdQtr'!G{row},'4thQtr'!G{row})"
            if val and val != normalize_formula(correct):
                record(errors, "Year to Date", row, "Forecasted", val, correct)

            # FORECAST DELTA
            val = normalize_formula(sheet[f"H{row}"].value)
            base = f"C{row}-G{row}"
            sum_qtr = f"SUM('1stQtr'!H{row},'2ndQtr'!H{row},'3rdQtr'!H{row},'4thQtr'!H{row})"
            if val and val not in [normalize_formula(base), normalize_formula(sum_qtr)]:
                record(errors, "Year to Date", row, "Forecast Delta", val, base)

            # FORECAST %
            val = normalize_formula(sheet[f"I{row}"].value)
            base = f"H{row}/G{row}"
            if val and base not in val:
                record(errors, "Year to Date", row, "Forecast %", val, base)

    # ---------------- QUARTER SHEETS VALIDATION ----------------
    for qtr, months in quarter_months.items():
        if qtr not in wb.sheetnames:
            continue
        sheet = wb[qtr]
        m1, m2, m3 = months
        for row in range(4, sheet.max_row + 1):
            # ACTUAL
            val = normalize_formula(sheet[f"C{row}"].value)
            correct = f"SUM({m1}!C{row},{m2}!C{row},{m3}!C{row})"
            if val and val != normalize_formula(correct):
                record(errors, qtr, row, "Actual", val, correct)

            # EXPECTED
            val = normalize_formula(sheet[f"D{row}"].value)
            correct = f"SUM({m1}!D{row},{m2}!D{row},{m3}!D{row})"
            if val and val != normalize_formula(correct):
                record(errors, qtr, row, "Expected", val, correct)

            # DELTA
            val = normalize_formula(sheet[f"E{row}"].value)
            base = f"C{row}-D{row}"
            sum_months = f"SUM({m1}!E{row},{m2}!E{row},{m3}!E{row})"
            if val and val not in [normalize_formula(base), normalize_formula(sum_months)]:
                record(errors, qtr, row, "Delta", val, base)

            # %
            val = normalize_formula(sheet[f"F{row}"].value)
            base = f"E{row}/D{row}"
            if val and base not in val:
                record(errors, qtr, row, "Expected %", val, base)

            # FORECASTED
            val = normalize_formula(sheet[f"G{row}"].value)
            correct = f"SUM({m1}!G{row},{m2}!G{row},{m3}!G{row})"
            if val and val != normalize_formula(correct):
                record(errors, qtr, row, "Forecasted", val, correct)

            # FORECAST DELTA
            val = normalize_formula(sheet[f"H{row}"].value)
            base = f"C{row}-G{row}"
            sum_months = f"SUM({m1}!H{row},{m2}!H{row},{m3}!H{row})"
            if val and val not in [normalize_formula(base), normalize_formula(sum_months)]:
                record(errors, qtr, row, "Forecast Delta", val, base)

            # %
            val = normalize_formula(sheet[f"I{row}"].value)
            base = f"H{row}/G{row}"
            if val and base not in val:
                record(errors, qtr, row, "Forecast %", val, base)

    return {
        "issues": errors,
        "formula_errors": len(errors)
    }